#!/usr/bin/env python3
"""
Agente de Standup Diario
Proyecto 5 — Platzi AI Course

Procesa el resumen del standup en lenguaje natural con Claude,
actualiza el Excel del proyecto y genera un resumen estructurado.

Uso:
    python standup.py                        # modo interactivo
    python standup.py --file resumen.txt     # desde archivo
    python standup.py --folder /ruta/carpeta # carpeta personalizada
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime

try:
    import anthropic
except ImportError:
    print("❌ Falta instalar: pip install anthropic")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("❌ Falta instalar: pip install openpyxl")
    sys.exit(1)


# ─── CONFIG ────────────────────────────────────────────────────────────────────

DEFAULT_FOLDER = Path(__file__).parent.parent / "SeguimientoProyectos"
SKIP_KEYWORDS  = ["plantilla"]
ESTADOS        = ["Pendiente", "En progreso", "Bloqueado", "Completado"]

# ─── EXCEL ─────────────────────────────────────────────────────────────────────

def find_projects(folder: Path) -> list[Path]:
    """Encuentra todos los .xlsx en la carpeta, excluyendo plantillas."""
    files = []
    for f in sorted(folder.glob("*.xlsx")):
        if not any(k in f.name.lower() for k in SKIP_KEYWORDS):
            files.append(f)
    return files


def read_project(path: Path) -> dict | None:
    """Lee el Excel y retorna la estructura del proyecto."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  ⚠️  No se pudo leer {path.name}: {e}")
        return None

    if "Tareas" not in wb.sheetnames:
        return None

    ws = wb["Tareas"]

    # Info del proyecto (fila 2)
    info_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    proyecto = str(info_row[1] or "").strip()
    if not proyecto:
        return None

    cliente  = str(info_row[3] or "").strip()
    pm       = str(info_row[5] or "").strip()
    inicio   = str(info_row[7] or "").strip()
    cierre   = str(info_row[9] or "").strip()

    # Tareas (desde fila 6, índice 6)
    tasks = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        task_id = str(row[0] or "").strip()
        if not task_id:
            break
        tasks.append({
            "id":          task_id,
            "fase":        str(row[1] or "").strip(),
            "categoria":   str(row[2] or "").strip(),
            "tarea":       str(row[3] or "").strip(),
            "responsable": str(row[4] or "").strip(),
            "fecha_ini":   str(row[5] or "").strip(),
            "fecha_fin":   str(row[6] or "").strip(),
            "prioridad":   str(row[7] or "").strip(),
            "estado":      str(row[8] or "").strip(),
            "avance":      str(row[9] or "0%").strip(),
            "notas":       str(row[10] or "").strip(),
        })

    # Bitácora (últimas notas por tarea)
    bitacora = []
    if "Bitácora" in wb.sheetnames:
        ws_b = wb["Bitácora"]
        for row in ws_b.iter_rows(min_row=2, values_only=True):
            if row[0]:
                bitacora.append({
                    "fecha":  str(row[0] or ""),
                    "taskId": str(row[1] or ""),
                    "tarea":  str(row[2] or ""),
                    "nota":   str(row[3] or ""),
                })

    return {
        "path":     path,
        "wb":       wb,
        "proyecto": proyecto,
        "cliente":  cliente,
        "pm":       pm,
        "inicio":   inicio,
        "cierre":   cierre,
        "tasks":    tasks,
        "bitacora": bitacora,
    }


def write_updates(project: dict, changes: list[dict]) -> bool:
    """Aplica los cambios al Excel y guarda."""
    wb = project["wb"]
    ws = wb["Tareas"]

    # Asegurar hoja Bitácora
    if "Bitácora" not in wb.sheetnames:
        ws_b = wb.create_sheet("Bitácora")
        ws_b.append(["Fecha", "ID Tarea", "Tarea", "Nota"])
    else:
        ws_b = wb["Bitácora"]

    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    for ch in changes:
        task_id = ch.get("id")
        # Find row in Tareas
        for row in ws.iter_rows(min_row=6):
            if str(row[0].value or "").strip() == task_id:
                if ch.get("estado"):
                    row[8].value = ch["estado"]
                if ch.get("avance") is not None:
                    row[9].value = ch["avance"]
                break

        # Write to Bitácora if there's a note
        nota = ch.get("nota", "").strip()
        if nota:
            tarea_name = next(
                (t["tarea"] for t in project["tasks"] if t["id"] == task_id), ""
            )
            ws_b.append([now, task_id, tarea_name, nota])

    try:
        wb.save(project["path"])
        return True
    except Exception as e:
        print(f"\n❌ Error al guardar: {e}")
        return False


# ─── CLAUDE ────────────────────────────────────────────────────────────────────

def build_system_prompt(project: dict) -> str:
    active_tasks = [t for t in project["tasks"] if t["estado"] != "Completado"]
    tasks_json   = json.dumps(active_tasks, ensure_ascii=False, indent=2)

    return f"""Eres un asistente de gerencia de proyectos especializado en procesar standups diarios.

Proyecto actual: {project['proyecto']} · Cliente: {project['cliente']}

Tareas activas del proyecto (en JSON):
{tasks_json}

Tu trabajo es analizar el resumen del standup que te dará el PM y extraer los cambios para cada tarea mencionada.

REGLAS:
1. Solo actualiza tareas que estén explícitamente mencionadas en el standup
2. Si una tarea fue mencionada pero le falta información clave (estado o avance), márcala como "necesita_info": true y explica qué falta
3. Si se menciona una tarea que no existe en la lista, repórtala en "tareas_nuevas"
4. El avance siempre como string con %, ejemplo "75%"
5. Los estados válidos son: {", ".join(ESTADOS)}
6. Si hay un bloqueo, extrae una descripción clara del problema para la bitácora
7. Sé conservador: si no estás seguro de un cambio, márcalo como necesita_info

Responde SIEMPRE en JSON con esta estructura exacta:
{{
  "cambios": [
    {{
      "id": "T-01",
      "tarea": "nombre de la tarea",
      "estado": "nuevo estado o null si no se mencionó",
      "avance": "75% o null si no se mencionó",
      "nota": "descripción del avance o bloqueo para la bitácora, o null",
      "necesita_info": false,
      "pregunta": null
    }}
  ],
  "tareas_nuevas": ["nombre de tarea no existente si se mencionó"],
  "resumen": "resumen ejecutivo del standup en 3-4 líneas",
  "estado_general": "ok|atencion|critico"
}}

Si una tarea necesita información, el campo "pregunta" debe contener la pregunta específica para el PM."""


def call_claude(client: anthropic.Anthropic, system: str, messages: list) -> str:
    """Llama a Claude y retorna el texto de la respuesta."""
    response = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=2048,
        system=system,
        messages=messages,
    )
    return response.content[0].text


def parse_json_response(text: str) -> dict | None:
    """Extrae el JSON de la respuesta de Claude."""
    try:
        # Try direct parse
        return json.loads(text)
    except json.JSONDecodeError:
        # Try to find JSON block
        import re
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            try:
                return json.loads(match.group())
            except json.JSONDecodeError:
                pass
    return None


# ─── UI ────────────────────────────────────────────────────────────────────────

def separator(char="─", width=55):
    print(char * width)


def print_header():
    separator("═")
    print("  🤖 AGENTE DE STANDUP DIARIO")
    print(f"  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    separator("═")
    print()


def select_project(projects: list[Path]) -> dict | None:
    """Muestra la lista y deja elegir al PM."""
    print("Proyectos disponibles:")
    valid = []
    for i, path in enumerate(projects, 1):
        p = read_project(path)
        if p:
            valid.append(p)
            print(f"  {i}. {p['proyecto']} · {p['cliente']}")
            print(f"     📄 {path.name}")

    if not valid:
        print("❌ No se encontraron proyectos válidos en la carpeta.")
        return None

    print()
    while True:
        try:
            choice = input("¿Qué proyecto corresponde al standup de hoy? (número): ").strip()
            idx = int(choice) - 1
            if 0 <= idx < len(valid):
                return valid[idx]
            print(f"  Por favor ingresa un número entre 1 y {len(valid)}")
        except (ValueError, KeyboardInterrupt):
            print("\nCancelado.")
            return None


def get_standup_text(args) -> str | None:
    """Obtiene el texto del standup — desde archivo o interactivo."""
    if args.file:
        try:
            text = Path(args.file).read_text(encoding="utf-8").strip()
            print(f"📄 Resumen cargado desde: {args.file}")
            print()
            return text
        except Exception as e:
            print(f"❌ No se pudo leer el archivo: {e}")
            return None

    # Modo interactivo
    print("¿Qué pasó hoy en el standup?")
    print("(Escribe el resumen — Enter en línea vacía para terminar)")
    print()
    lines = []
    try:
        while True:
            line = input("> ")
            if line == "" and lines:
                break
            lines.append(line)
    except KeyboardInterrupt:
        print("\nCancelado.")
        return None

    return "\n".join(lines).strip()


def show_active_tasks(project: dict):
    """Muestra las tareas activas del proyecto."""
    active = [t for t in project["tasks"] if t["estado"] != "Completado"]
    print(f"\nTareas activas — {project['proyecto']}:")
    separator()
    for t in active:
        estado_icon = {"En progreso": "🔄", "Bloqueado": "🔴", "Pendiente": "⏸️"}.get(t["estado"], "•")
        print(f"  {estado_icon} {t['id']} · {t['tarea']}")
        print(f"     {t['estado']} ({t['avance']}) — {t['responsable']}")
    separator()
    print()


def display_changes(cambios: list[dict], project: dict) -> list[dict]:
    """Muestra los cambios detectados y retorna solo los que tienen info completa."""
    print("\nCambios detectados:")
    separator()

    ready    = []
    pending  = []

    for ch in cambios:
        if ch.get("necesita_info"):
            pending.append(ch)
            continue

        estado = ch.get("estado") or "sin cambio de estado"
        avance = ch.get("avance") or "sin cambio de avance"
        icon   = "✅" if estado == "Completado" else ("🔴" if estado == "Bloqueado" else "🔄")

        print(f"  {icon} {ch['id']} · {ch['tarea']}")
        print(f"     → {estado} ({avance})")
        if ch.get("nota"):
            print(f"     📝 {ch['nota']}")
        ready.append(ch)

    separator()
    return ready, pending


def resolve_pending(client: anthropic.Anthropic, system: str,
                    messages: list, pending: list[dict], ready: list[dict]) -> list[dict]:
    """Pregunta al PM por las tareas con información incompleta."""
    for ch in pending:
        print(f"\n⚠️  Necesito información sobre: {ch['id']} · {ch['tarea']}")
        print(f"   → {ch.get('pregunta', '¿Cuál es el estado actual de esta tarea?')}")
        print()

        try:
            answer = input("Tu respuesta: ").strip()
        except KeyboardInterrupt:
            print("\nOmitiendo esta tarea.")
            continue

        if not answer:
            print("   Omitida.")
            continue

        # Ask Claude to resolve with the PM's answer
        messages_copy = messages + [
            {"role": "assistant", "content": json.dumps({"cambios": [ch]}, ensure_ascii=False)},
            {"role": "user",      "content": f"Respuesta del PM sobre {ch['id']}: {answer}. Actualiza solo ese cambio con la información recibida."},
        ]

        try:
            raw      = call_claude(client, system, messages_copy)
            resolved = parse_json_response(raw)
            if resolved and resolved.get("cambios"):
                updated = resolved["cambios"][0]
                updated["necesita_info"] = False
                icon = "✅" if updated.get("estado") == "Completado" else "🔄"
                print(f"  {icon} {updated['id']} → {updated.get('estado','?')} ({updated.get('avance','?')})")
                ready.append(updated)
        except Exception as e:
            print(f"   ⚠️  No se pudo resolver: {e}")

    return ready


def print_summary(resumen: str, estado_general: str, project: dict):
    """Imprime el resumen del standup."""
    icon = {"ok": "🟢", "atencion": "🟡", "critico": "🔴"}.get(estado_general, "⚪")
    fecha = datetime.now().strftime("%d/%m/%Y")

    print()
    separator("─")
    print(f"  📋 RESUMEN DEL STANDUP — {fecha}")
    separator("─")
    print(f"  Proyecto: {project['proyecto']} · {project['cliente']}")
    print(f"  Estado general: {icon}")
    print()
    print(resumen)
    separator("─")


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Agente de Standup Diario")
    parser.add_argument("--file",   help="Archivo .txt con el resumen del standup")
    parser.add_argument("--folder", help="Ruta a la carpeta SeguimientoProyectos")
    args = parser.parse_args()

    # Verificar API key
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("❌ Variable ANTHROPIC_API_KEY no encontrada.")
        print("   Configúrala con: export ANTHROPIC_API_KEY='sk-ant-...'")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    # Carpeta de proyectos
    folder = Path(args.folder) if args.folder else DEFAULT_FOLDER
    if not folder.exists():
        print(f"❌ Carpeta no encontrada: {folder}")
        print(f"   Usa --folder para especificar la ruta correcta.")
        sys.exit(1)

    print_header()

    # 1. Seleccionar proyecto
    projects = find_projects(folder)
    if not projects:
        print(f"❌ No se encontraron archivos .xlsx en: {folder}")
        sys.exit(1)

    project = select_project(projects)
    if not project:
        sys.exit(0)

    # 2. Mostrar tareas activas
    show_active_tasks(project)

    # 3. Obtener resumen del standup
    standup_text = get_standup_text(args)
    if not standup_text:
        sys.exit(0)

    # 4. Llamar a Claude
    print("\n🧠 Analizando con Claude...")
    system   = build_system_prompt(project)
    messages = [{"role": "user", "content": standup_text}]

    try:
        raw      = call_claude(client, system, messages)
        parsed   = parse_json_response(raw)
    except Exception as e:
        print(f"❌ Error al llamar a Claude: {e}")
        sys.exit(1)

    if not parsed:
        print("❌ Claude no devolvió una respuesta válida.")
        print("Respuesta recibida:")
        print(raw)
        sys.exit(1)

    cambios         = parsed.get("cambios", [])
    tareas_nuevas   = parsed.get("tareas_nuevas", [])
    resumen         = parsed.get("resumen", "")
    estado_general  = parsed.get("estado_general", "ok")

    if not cambios:
        print("\n⚠️  Claude no detectó cambios en las tareas del proyecto.")
        print("   Verifica que el resumen mencione tareas del proyecto seleccionado.")
        sys.exit(0)

    # 5. Mostrar cambios y resolver pendientes
    ready, pending = display_changes(cambios, project)

    if tareas_nuevas:
        print(f"\n⚠️  Tareas mencionadas que no existen en el Excel:")
        for t in tareas_nuevas:
            print(f"   • {t}")
        print("   (No se agregarán automáticamente — usa el sistema de prompts para crearlas)")

    if pending:
        print(f"\n❓ Tengo {len(pending)} pregunta(s) antes de continuar:")
        ready = resolve_pending(client, system, messages, pending, ready)

    if not ready:
        print("\nNo hay cambios para aplicar.")
        sys.exit(0)

    # 6. Confirmación
    print(f"\n¿Confirmas estos {len(ready)} cambio(s) en el Excel? (s/n): ", end="")
    try:
        confirm = input().strip().lower()
    except KeyboardInterrupt:
        print("\nCancelado.")
        sys.exit(0)

    if confirm != "s":
        print("Cambios descartados.")
        sys.exit(0)

    # 7. Escribir al Excel
    if write_updates(project, ready):
        print(f"\n✅ Excel actualizado: {project['path'].name}")
    else:
        print("\n❌ No se pudieron guardar los cambios.")
        sys.exit(1)

    # 8. Resumen final
    if resumen:
        print_summary(resumen, estado_general, project)

    print("\n✅ Standup procesado correctamente.\n")


if __name__ == "__main__":
    main()
